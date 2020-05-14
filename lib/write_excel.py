import openpyxl
from conf.settings import FILE_PATH
from conf.settings import ActualResults
from conf.settings import SHEET_NAME
from conf.settings import Result


class Write:

    def __init__(self, row, col=ActualResults, value=""):
        self.excel_path = FILE_PATH
        self.sheet_name = SHEET_NAME
        self.row = row
        self.col = col
        self.value = value
        self.write_excel()

    def write_excel(self):
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook.get_sheet_by_name(self.sheet_name)
        if self.col == ActualResults:
            sheet.cell(row=self.row, column=self.col, value=str(self.value))
            sheet.cell(row=self.row, column=Result, value="")
        else:
            sheet.cell(row=self.row, column=self.col, value=str(self.value))
            sheet.cell(row=self.row, column=ActualResults, value="")
        workbook.save(self.excel_path)


if __name__ == '__main__':
    Write(value="pass", row=int(4), col=16)
    Write(row=2)
